from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl import load_workbook
import datetime
# import datetime
import pandas as pd
from datetime import datetime
# 读取Excel文件
excel_file = '总计表.xlsx'
# df_excel = pd.read_excel(excel_file)
df_excel = pd.read_excel(excel_file, engine='openpyxl')

# 显示数据框的前几行
print("Excel文件数据：")
# print(df_excel.head())

# # 选择特定范围的数据
# selected_data = df_excel.iloc[1:45, 1]  # 选择第2列（B列）的第2行到第45行的数据

# # 遍历DataFrame中的值
# for row in df_excel.itertuples():
#     for value in row[1:]:  # 略过第一个元素，因为第一个元素是行索引
#         print(value)
def hk1():
    # 提取时间和值
    ## 提取时间和值
    times = df_excel.iloc[0:45, 4].tolist()  # 提取B2到B45的时间并转换为列表
    values = df_excel.iloc[0:45, 5].tolist()  # 提取C2到C45的值并转换为列表

    # times = [times[i] for i in non_empty_indices]
    # values = [values[i] for i in non_empty_indices]


    # 过滤掉空值
    non_empty_indices = [i for i, time in enumerate(times) if pd.notna(time)]
    times = [times[i] for i in non_empty_indices]
    values = [values[i] for i in non_empty_indices]
    # print(f" times {times} values {values}")
    # 预处理时间数据并与对应的值组合成元组列表
    processed_data = []
    for t, v in zip(times, values):
        if isinstance(t, str):
            try:
                t = t.replace("/", "-").replace(".", "-")
                processed_time = datetime.strptime(t, '%Y-%m-%d')
                processed_data.append((processed_time, v*-1))
            except ValueError:
                # 解析失败，尝试切片操作
                try:
                    year = int(t[:4])
                    month = int(t[4:6])
                    day = int(t[6:])
                    processed_time = datetime(year, month, day)
                    processed_data.append((processed_time, v*-1))
                    # print(f" processed_time {processed_time} {t}")
                except ValueError:
                    print(f"无法解析的时间格式: {t}")
        else:
            if isinstance(t, datetime):
                processed_data.append((t, v * -1))
            elif isinstance(t, int):
                t = str(t)
                year = int(t[:4])
                month = int(t[4:6])
                day = int(t[6:])
                processed_time = datetime(year, month, day)
                processed_data.append((processed_time, v * -1))
                # print(f" processed_time {processed_time} {t}")
            else:
                print(f"t {t}  v {v}  {type(t)} ")
    print(processed_data)
    # 根据时间进行排序
    sorted_data = sorted(processed_data, key=lambda x: x[0])

    # 打印排序后的时间和对应的值
    for time, value in sorted_data:
        print(time, value)

    print(len(sorted_data), type(sorted_data))
    return sorted_data

def hk2():
    # 提取时间和值
    ## 提取时间和值
    times = df_excel.iloc[0:45, 1].tolist()  # 提取B2到B45的时间并转换为列表
    values = df_excel.iloc[0:45, 2].tolist()  # 提取C2到C45的值并转换为列表

    # times = [times[i] for i in non_empty_indices]
    # values = [values[i] for i in non_empty_indices]

    # 过滤掉空值
    non_empty_indices = [i for i, time in enumerate(times) if pd.notna(time)]
    times = [times[i] for i in non_empty_indices]
    values = [values[i] for i in non_empty_indices]

    # 预处理时间数据并与对应的值组合成元组列表
    processed_data = []
    for t, v in zip(times, values):
        if isinstance(t, str):
            try:
                t = t.replace("/", "-").replace(".", "-")
                processed_time = datetime.strptime(t, '%Y-%m-%d')
                processed_data.append((processed_time, v * -1))
            except ValueError:
                # 解析失败，尝试切片操作
                try:
                    year = int(t[:4])
                    month = int(t[4:6])
                    day = int(t[6:])
                    processed_time = datetime(year, month, day)
                    processed_data.append((processed_time, v * -1))
                    # print(f" processed_time {processed_time} {t}")
                except ValueError:
                    print(f"无法解析的时间格式: {t}")
        else:
            processed_data.append((t, v * -1))
    print(processed_data)
    # 根据时间进行排序
    sorted_data = sorted(processed_data, key=lambda x: x[0])

    # 打印排序后的时间和对应的值
    for time, value in sorted_data:
        print(time, value)

    print(len(sorted_data), type(sorted_data))
    return sorted_data






# hk1()
# hk2()

def zji1():
    dataq = hk1()
    dataw = hk2()
    z = 0
    for i in dataq:
        z += i[1]
    print(z)
    z = 0
    for i in dataw:
        z += i[1]
    print(z)

    sorted_data = dataq + dataw

    # 根据时间进行排序
    sorted_data = sorted(sorted_data, key=lambda x: x[0])
    print(sorted_data)
    z = 0
    for i in sorted_data:
        z += i[1]
        # print(str(i[0])[:11])
        # print(int(i[1]))
    print(z)

    # 定义利息率
    interest_rate = 0.35 / 1000  # 每天0.35元
    # 初始总额和利息
    total_amount = 0
    interest = 0
    # 初始日期
    start_date = datetime(2021, 11, 2)

    # 新的列表，包括总额和利息
    new_data = []
    # new_data.append((total_amount, interest))

    for date, amount in sorted_data:

        # 计算时间差
        days_diff = (date - start_date).days
        # 更新总额
        total_amount += amount
        # 计算利息
        interest = round(total_amount * interest_rate * days_diff , 2)
        # 将总额和利息添加到新列表中
        new_data.append((date,amount,total_amount, interest))
        # 更新起始日期
        start_date = date

        # print(f" date {date} amount {amount}  days_diff {days_diff}  interest {interest} ")

    # 输出新的列表
    for row in new_data:
        print(row)

    # df_data = pd.DataFrame(new_data, columns=['Date', 'Total Amount', 'Interest', 'Value'])
    # # 要写入的 Excel 文件名
    # excel_file = '总计表.xlsx'
    #
    # # 读取现有的 Excel 文件
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
    #     # 将 DataFrame 写入 Excel 文件，从 D65 单元格开始
    #     df_data.to_excel(writer, index=False, sheet_name='Sheet1', startrow=64, startcol=3)
    #
    # print("Data has been written to", excel_file)

    # 创建 DataFrame
    # 创建 DataFrame
    df_data = pd.DataFrame(new_data, columns=['Date', 'Total Amount', 'Interest', 'Value'])

    # 要写入的 Excel 文件名
    excel_file = '总计表.xlsx'

    # 读取现有的 Excel 文件
    try:
        # 尝试加载现有的 Excel 文件
        wb = load_workbook(excel_file)

        # 如果文件中没有工作表，则创建一个新的工作表
        if not wb.sheetnames:
            ws = wb.create_sheet(title='Sheet1')
        else:
            # 选择第一个工作表
            ws = wb[wb.sheetnames[0]]

        # 在工作表中追加写入数据
        for row in new_data:
            ws.append(row)

        # 保存修改后的 Excel 文件
        wb.save(excel_file)

        print("Data has been written to", excel_file)
    except FileNotFoundError:
        # 如果文件不存在，则直接写入数据到新文件中
        df_data.to_excel(excel_file, index=False)
        print("Data has been written to a new file:", excel_file)

def zji2():
    dataq = hk1()
    dataw = hk2()
    z = 0
    for i in dataq:
        z += i[1]
    print(z)
    z = 0
    for i in dataw:
        z += i[1]
    print(z)

    sorted_data = dataq + dataw

    # 根据时间进行排序
    sorted_data = sorted(sorted_data, key=lambda x: x[0])
    print(sorted_data)
    z = 0
    for i in sorted_data:
        z += i[1]
        # print(str(i[0])[:11])
        # print(int(i[1]))
    print(z)

    # 定义利息率
    interest_rate = 0.35 / 1000  # 每天0.35元
    # 初始总额和利息
    total_amount = 0
    # interest = 0
    # 初始日期
    start_date = datetime(2021, 11, 2)

    # 新的列表，包括总额和利息
    new_data = []
    # new_data.append((total_amount, interest))

    zg = 0

    for date, amount in sorted_data:

        # 计算时间差
        days_diff = (date - start_date).days
        # 更新总额
        total_amount += amount
        zg += amount
        # 计算利息
        interest = round(total_amount * interest_rate * days_diff , 2)

        dcfl = round(zg * interest_rate * days_diff, 2)
        zg += dcfl

        # 将总额和利息添加到新列表中
        new_data.append((date,amount,total_amount, interest,zg,dcfl,days_diff))
        # 更新起始日期
        start_date = date

        # print(f" date {date} amount {amount}  days_diff {days_diff}  interest {interest} ")

    # 输出新的列表
    for row in new_data:
        print(row)

    # df_data = pd.DataFrame(new_data, columns=['Date', 'Total Amount', 'Interest', 'Value'])
    # # 要写入的 Excel 文件名
    # excel_file = '总计表.xlsx'
    #
    # # 读取现有的 Excel 文件
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
    #     # 将 DataFrame 写入 Excel 文件，从 D65 单元格开始
    #     df_data.to_excel(writer, index=False, sheet_name='Sheet1', startrow=64, startcol=3)
    #
    # print("Data has been written to", excel_file)

    # 创建 DataFrame
    # 创建 DataFrame
    df_data = pd.DataFrame(new_data, columns=['Date', 'Total Amount', 'Interest', 'Value',"ze","lx","days_diff"])

    # 要写入的 Excel 文件名
    excel_file = '总计表.xlsx'

    try:
        # 尝试加载现有的 Excel 文件
        wb = load_workbook(excel_file)

        # 如果文件中没有工作表，则创建一个新的工作表
        if not wb.sheetnames:
            ws = wb.create_sheet(title='Sheet1')
        else:
            # 选择第一个工作表
            ws = wb[wb.sheetnames[0]]

        # 获取 DataFrame 的数据行
        rows = list(dataframe_to_rows(df_data, index=False))

        # 写入数据到特定位置
        for r_idx, row in enumerate(rows, start=50):  # 从第65行开始写入
            for c_idx, value in enumerate(row, start=3):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 保存修改后的 Excel 文件
        wb.save(excel_file)

        print("Data has been written to", excel_file)
    except FileNotFoundError:
        # 如果文件不存在，则直接写入数据到新文件中
        df_data.to_excel(excel_file, index=False, startrow=50)  # 从第65行开始写入
        print("Data has been written to a new file:", excel_file)



zji2()
